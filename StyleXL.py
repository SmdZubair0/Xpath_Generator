from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class StyleXL:
    def __init__(self,filepath, sheet = None):
        self.filepath = filepath
        self.wb = load_workbook(self.filepath)
        self.sheets = [sheet]
        if sheet is None:
            self.sheets = []
            for sheet in self.wb.sheetnames:
                self.sheets.append(sheet)

    def color_header(self, color = '0000FF', font_color = 'FFFFFF'):
        header_font = Font(bold=True, color=font_color)
        header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        for sheet in self.sheets:
            sheet = self.wb[sheet]
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)

    def column_alignment(self, horizontal='left', vertical='top', wrap_text=True):
        alignment = Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap_text)
        for sheet in self.sheets:
            sheet = self.wb[sheet]

            row_max_length = {}

            max_length_allowed = 100
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Get the column letter (e.g., A, B)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                        max_length = min(max_length, max_length_allowed)
                        row_max_length[cell.row] = max(row_max_length.get(cell.row, 0), len(str(cell.value)))

                sheet.column_dimensions[col_letter].width = max_length + 2
                cell.alignment = alignment

            for row, max_length in row_max_length.items():
                sheet.row_dimensions[row].height = ((max_length - 1) / max_length_allowed + 1) * 12

            header = True
            for row in sheet.iter_rows():  
                if header:
                    header = False
                    continue
                for cell in row:
                    if cell.value: 
                        cell.alignment = alignment


    def border(self):
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for sheet in self.sheets:
            sheet = self.wb[sheet]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

    def color_status(self):
        for sheet in self.sheets:
            sheet = self.wb[sheet]
            for row in sheet.iter_rows():
                for cell in row: 
                    if cell.value == 'Passed':
                        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                        cell.font = Font(color='000000')
                    elif cell.value == 'Failed':
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        cell.font = Font(color='FFFFFF')
                    elif cell.value == 'Pending':
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        cell.font = Font(color='000000')

    def save_file(self,filepath):
        self.wb.save(filepath)